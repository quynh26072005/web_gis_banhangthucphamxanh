"""
Utilities for exporting inventory data to PDF and Excel
"""
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_stock_transactions_pdf(transactions, farm_name=""):
    """
    Xuất danh sách giao dịch kho ra file PDF
    """
    buffer = BytesIO()
    
    # Tạo PDF với landscape orientation
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=50,
        bottomMargin=30
    )
    
    # Container cho các elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#28a745'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Title
    title_text = f"BÁO CÁO XUẤT NHẬP KHO"
    if farm_name:
        title_text += f"<br/>{farm_name}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    
    # Thông tin báo cáo
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_LEFT
    )
    info_text = f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
    info_text += f"Tổng số giao dịch: {transactions.count()}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))
    
    # Tạo bảng dữ liệu
    data = [['STT', 'Ngày', 'Loại', 'Sản phẩm', 'Số lượng', 'Đơn vị', 'Nhà cung cấp', 'Ghi chú']]
    
    for idx, trans in enumerate(transactions, 1):
        transaction_type = 'Nhập' if trans.transaction_type == 'IN' else 'Xuất'
        supplier_name = trans.supplier.name if trans.supplier else '-'
        
        data.append([
            str(idx),
            trans.created_at.strftime('%d/%m/%Y'),
            transaction_type,
            trans.product.name[:30],
            str(trans.quantity),
            trans.product.unit,
            supplier_name[:20],
            (trans.notes or '-')[:30]
        ])
    
    # Tạo table
    table = Table(data, colWidths=[0.5*inch, 1*inch, 0.8*inch, 2*inch, 0.8*inch, 0.8*inch, 1.5*inch, 2*inch])
    
    # Style cho table
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # STT
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Ngày
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Loại
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Số lượng
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer = Paragraph(
        "Clean Food GIS - Hệ thống quản lý thực phẩm sạch<br/>© 2026 All rights reserved",
        footer_style
    )
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf


def export_stock_transactions_excel(transactions, farm_name=""):
    """
    Xuất danh sách giao dịch kho ra file Excel
    """
    # Tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xuất Nhập Kho"
    
    # Styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(name='Arial', size=16, bold=True, color='28a745')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_text = f"BÁO CÁO XUẤT NHẬP KHO"
    if farm_name:
        title_text += f" - {farm_name}"
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Info
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Tổng: {transactions.count()} giao dịch"
    info_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['STT', 'Ngày giờ', 'Loại giao dịch', 'Sản phẩm', 'Số lượng', 'Đơn vị', 'Nhà cung cấp', 'Ghi chú']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for idx, trans in enumerate(transactions, 1):
        row_num = idx + 4
        
        transaction_type = 'Nhập kho' if trans.transaction_type == 'IN' else 'Xuất kho'
        supplier_name = trans.supplier.name if trans.supplier else '-'
        
        data_row = [
            idx,
            trans.created_at.strftime('%d/%m/%Y %H:%M'),
            transaction_type,
            trans.product.name,
            trans.quantity,
            trans.product.unit,
            supplier_name,
            trans.notes or '-'
        ]
        
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Alignment
            if col_num in [1, 2, 3, 6]:  # STT, Ngày, Loại, Đơn vị
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_num == 5:  # Số lượng
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Color coding for transaction type
            if col_num == 3:
                if trans.transaction_type == 'IN':
                    cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                    cell.font = Font(color='155724', bold=True)
                else:
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                    cell.font = Font(color='721c24', bold=True)
    
    # Adjust column widths
    column_widths = [8, 18, 15, 30, 12, 10, 25, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    excel_data = buffer.getvalue()
    buffer.close()
    
    return excel_data


def export_inventory_report_pdf(products, farm_name=""):
    """
    Xuất báo cáo tồn kho ra PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#28a745'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    title_text = f"BÁO CÁO TỒN KHO"
    if farm_name:
        title_text += f"<br/>{farm_name}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    
    # Info
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT)
    info_text = f"Ngày báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
    info_text += f"Tổng số sản phẩm: {products.count()}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))
    
    # Table
    data = [['STT', 'Sản phẩm', 'Danh mục', 'Tồn kho', 'Đơn vị', 'Giá', 'Trạng thái']]
    
    total_value = 0
    for idx, product in enumerate(products, 1):
        status = '⚠️ Sắp hết' if product.stock_quantity < 20 else '✓ Đủ hàng'
        value = product.stock_quantity * float(product.price)
        total_value += value
        
        data.append([
            str(idx),
            product.name[:30],
            product.category.name if product.category else '-',
            str(product.stock_quantity),
            product.unit,
            f"{product.price:,.0f}đ",
            status
        ])
    
    # Add total row
    data.append(['', '', '', '', 'TỔNG GIÁ TRỊ:', f"{total_value:,.0f}đ", ''])
    
    table = Table(data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3cd')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 30))
    
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    footer = Paragraph("Clean Food GIS - Hệ thống quản lý thực phẩm sạch<br/>© 2026", footer_style)
    elements.append(footer)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf


def export_inventory_report_excel(products, farm_name=""):
    """
    Xuất báo cáo tồn kho ra Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo Tồn Kho"
    
    # Styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(name='Arial', size=16, bold=True, color='28a745')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_text = f"BÁO CÁO TỒN KHO"
    if farm_name:
        title_text += f" - {farm_name}"
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Info
    ws.merge_cells('A2:G2')
    info_cell = ws['A2']
    info_cell.value = f"Ngày báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Tổng: {products.count()} sản phẩm"
    info_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['STT', 'Sản phẩm', 'Danh mục', 'Tồn kho', 'Đơn vị', 'Giá (VNĐ)', 'Trạng thái']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    total_value = 0
    for idx, product in enumerate(products, 1):
        row_num = idx + 4
        status = 'Sắp hết' if product.stock_quantity < 20 else 'Đủ hàng'
        value = product.stock_quantity * float(product.price)
        total_value += value
        
        data_row = [
            idx,
            product.name,
            product.category.name if product.category else '-',
            product.stock_quantity,
            product.unit,
            float(product.price),
            status
        ]
        
        for col_num, value_cell in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value_cell
            cell.border = border
            
            if col_num in [1, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_num == 6:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Status color
            if col_num == 7:
                if product.stock_quantity < 20:
                    cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
                    cell.font = Font(color='856404', bold=True)
                else:
                    cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                    cell.font = Font(color='155724', bold=True)
    
    # Total row
    total_row = products.count() + 5
    ws.merge_cells(f'A{total_row}:E{total_row}')
    total_cell = ws[f'A{total_row}']
    total_cell.value = 'TỔNG GIÁ TRỊ TỒN KHO:'
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.border = border
    
    value_cell = ws[f'F{total_row}']
    value_cell.value = total_value
    value_cell.font = Font(bold=True, color='FF0000', size=12)
    value_cell.alignment = Alignment(horizontal='right', vertical='center')
    value_cell.number_format = '#,##0'
    value_cell.border = border
    value_cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
    
    # Adjust column widths
    column_widths = [8, 35, 20, 12, 10, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    buffer = BytesIO()
    wb.save(buffer)
    excel_data = buffer.getvalue()
    buffer.close()
    
    return excel_data
